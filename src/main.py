import pandas as pd
import numpy as np
import sqlite3
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import (
                        PatternFill, Border, Side,
                        Alignment, Font)
from openpyxl.drawing.image import Image
import seaborn as sns
from matplotlib.colors import LinearSegmentedColormap

def db_getdata(select_string):
    conn = sqlite3.connect("shops.db")
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master where type= 'table'")
    dataframe = pd.read_sql(select_string, conn)
    conn.close()
    return dataframe

def prepare_data(df):
    df['sold_amount'] = df['sold_amount'].str.replace(',', '.').astype(float)
    df['sold_ss'] = df['sold_ss'].str.replace(',', '.').astype(float)
    df['date'] = df['date'].astype("datetime64[ns]")
    df['stonks'] = df['sold_amount'] - df['sold_ss']
    df = df.fillna(0)
    return df


def group_by_store_accum_hist(df):
    stonks_grouped_by_years = df.groupby([df['date'].dt.year, df["id"]])["stonks"].sum().reset_index()
    stonks_2014 = stonks_grouped_by_years[stonks_grouped_by_years['date'] == 2014]
    stonks_2015 = stonks_grouped_by_years[stonks_grouped_by_years['date'] == 2015].reset_index(drop=True)
    _, ax = plt.subplots(figsize=(15, 5))
    ax.bar(stonks_2014.index, stonks_2014["stonks"], label='2014')
    ax.bar(stonks_2015.index, stonks_2015['stonks'], bottom=stonks_2014["stonks"], label='2015')
    plt.xticks(stonks_2014.index, stonks_2014['id'])
    ax.set_title('Прибыль по магазинам')
    ax.set_xlabel("ID магазина")
    ax.set_ylabel("млн руб")
    plt.legend()
    # plt.show()

    plot_name = 'Accum_hist.png'
    plt.savefig(plot_name)
    return plot_name

def group_by_art_id_bad_goods(df):
    # ГРУППИРУЕМ ПО АРТИКЛЮ ТОВАРОВ
    group_by_art = df.groupby('good_id')

    top_10_articles_by_stonks = group_by_art.stonks.sum().sort_values(ascending=True).head(
        10).reset_index() # топ-10 убыточных товаров
    fig, ax = plt.subplots(figsize=(12, 4))
    ax.bar(top_10_articles_by_stonks.index, top_10_articles_by_stonks['stonks'], color='r')
    ax.set_title('ТОП-10 убыточных товаров')
    ax.set_xlabel("артикуль товара")
    ax.set_ylabel("млн руб")
    plt.xticks(top_10_articles_by_stonks.index, top_10_articles_by_stonks['good_id'])
    # plt.show()
    plot_name = 'Bad_goods_bar.png'
    fig.savefig(plot_name)
    return plot_name


def group_by_art_id_top_goods(df):
    # ГРУППИРУЕМ ПО АРТИКЛЮ ТОВАРОВ
    group_by_art = df.groupby('good_id')

    top_10_articles_by_stonks = group_by_art.stonks.sum().sort_values(ascending=False).head(
        10)  # топ-10 прибыльных товаров
    top_10_articles_by_stonks.plot(
        kind='bar',
        color='g',
        title=f'ТОП-10 прибыльных товаров',
        xlabel='артикуль товара',
        ylabel='млн руб',
        rot=0,
        figsize=(12, 4),
    )

    plt.show()
    plot_name = 'Stonks_hist.png'
    plt.savefig(plot_name)
    return plot_name


def group_by_store_id_pie(df):
    df['share'] = df['sold_count'] / df['sold_count'].sum()
    grouped = df.groupby('id').share.sum().sort_values(ascending=False).reset_index()


    fig, ax = plt.subplots()
    colors = LinearSegmentedColormap.from_list(sns.color_palette("YlOrBr", as_cmap=True), ['#ed0e00', '#ffffff'], N=16)
    ax.pie(grouped['share'], labels=grouped['id'], autopct='%1.1f%%', explode=[0.05] * len(grouped['id']), textprops={'fontsize': 8}, colors=colors(np.linspace(0, 1, 16)))
    ax.set_title("Доли продаж (шт) по ТТ")
    plot_name = 'Pie_plot.png'
    fig.savefig(plot_name)
    plt.show()
    return plot_name


def group_by_date(df):
    # Выручка ПО ДНЯМ И ГОДАМ
    revenue_grouped_by_day = df.groupby([df['date'].dt.year, df['date'].dt.day]).sold_amount.sum()
    fig, ax = plt.subplots(figsize=(6.63,6.68))
    revenue_grouped_by_day.loc[2014].plot(kind='line', style='-o')
    revenue_grouped_by_day.loc[2015].plot(kind='line', style='--o')
    plt.xticks(range(1, 8), [f'{i}.04' for i in range(1, 8)])
    ax.set_title("Продажи (руб) по всем магазинам")
    ax.set_xlabel("Дата")
    ax.set_ylabel("млн руб")
    ax.legend(["2014", "2015"])
    # plot_name = 'Revenue_line_without_bg.png'
    # plt.savefig(plot_name)
    bg_image = plt.imread('logo.png')
    fig.figimage(bg_image, 84, 74, alpha=0.2)
    # plt.show()
    plot_name = 'Revenue_line_with_bg.png'
    plt.savefig(plot_name)
    return plot_name

def load_to_excel_sheet(df, plot_name_hist, plot_name_line, plot_name_acc_hist, plot_name_hist_bad_goods, plot_name_pie):

    wb = load_workbook('Отчет.xlsx')
    sheet = wb['аналитика']

    font = Font(
        name='Calibri',
        size=11,
        bold=True
    )

    bold_border = Border(
        left=Side(border_style='medium', color='FF000000'),
        right=Side(border_style='medium', color='FF000000'),
        top=Side(border_style='medium', color='FF000000'),
        bottom=Side(border_style='medium', color='FF000000')
    )
    thin_border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000')
    )

    stores = sorted(df['id'].unique())
    for i in range(4, 4 + len(stores)):
        # СТОЛБЕЦ B
        sheet[f'B{i}'].font = font
        sheet[f'B{i}'].border = bold_border
        sheet[f'B{i}'].fill = PatternFill(fgColor='00FFFFFF')
        sheet[f'B{i}'].alignment = Alignment(horizontal='center')
        sheet[f'B{i}'].number_format = "0"
        sheet[f'B{i}'].value = stores[i - 4]

        # СТОЛБЕЦ C
        sheet[f'C{i}'].number_format = "0" # UILTIN_FORMATS[2] == "0.00"
        sheet[f'C{i}'].value = df['sold_count'][df['id'] == stores[i - 4]].sum()
        sheet[f'C{i}'].font = Font(name='Calibri',size=11)
        sheet[f'C{i}'].fill = PatternFill(fgColor='00FFFFFF')
        sheet[f'C{i}'].border = thin_border
        sheet[f'C{i}'].alignment = Alignment(horizontal='center')

        # СТОЛБЕЦ D
        sheet[f'D{i}'].number_format = "0.0"  # UILTIN_FORMATS[2] == "0.00"
        sheet[f'D{i}'].value = (df['sold_amount'][df['id'] == stores[i - 4]].sum())
        sheet[f'D{i}'].font = Font(name='Calibri', size=11)
        sheet[f'D{i}'].fill = PatternFill(fgColor='00FFFFFF')
        sheet[f'D{i}'].border = thin_border
        sheet[f'D{i}'].alignment = Alignment(horizontal='center')

        # СТОЛБЕЦ E
        sheet[f'E{i}'].number_format = "0.0"  # UILTIN_FORMATS[2] == "0.00"
        sheet[f'E{i}'].value = (df['sold_ss'][df['id'] == stores[i - 4]].sum())
        sheet[f'E{i}'].font = Font(name='Calibri', size=11)
        sheet[f'E{i}'].fill = PatternFill(fgColor='00FFFFFF')
        sheet[f'E{i}'].border = thin_border
        sheet[f'E{i}'].alignment = Alignment(horizontal='center')

        # СТОЛБЕЦ F
        sheet[f'F{i}'].number_format = "0.0"  # UILTIN_FORMATS[2] == "0.00"
        sheet[f'F{i}'].value = (df['stonks'][df['id'] == stores[i - 4]].sum())
        sheet[f'F{i}'].font = Font(name='Calibri', size=11)
        sheet[f'F{i}'].fill = PatternFill(fgColor='00FFFFFF')
        sheet[f'F{i}'].border = thin_border
        sheet[f'F{i}'].alignment = Alignment(horizontal='center')

    # строка ИТОГО
    sheet[f'B{len(stores) + 4}'].value = 'Итого'
    sheet[f'B{len(stores) + 4}'].font = font
    sheet[f'B{len(stores) + 4}'].fill = PatternFill(fill_type='solid', fgColor='00C0C0C0')
    sheet[f'B{len(stores) + 4}'].alignment = Alignment(horizontal='center')
    sheet[f'B{len(stores) + 4}'].border = bold_border

    sheet[f'C{len(stores) + 4}'].value = int(df['sold_count'].sum())
    sheet[f'C{len(stores) + 4}'].font = font
    sheet[f'C{len(stores) + 4}'].fill = PatternFill(fill_type='solid', fgColor='00C0C0C0')
    sheet[f'C{len(stores) + 4}'].alignment = Alignment(horizontal='center')
    sheet[f'C{len(stores) + 4}'].border = bold_border

    sheet[f'D{len(stores) + 4}'].value = df['sold_amount'].sum()
    sheet[f'D{len(stores) + 4}'].font = font
    sheet[f'D{len(stores) + 4}'].fill = PatternFill(fill_type='solid', fgColor='00C0C0C0')
    sheet[f'D{len(stores) + 4}'].alignment = Alignment(horizontal='center')
    sheet[f'D{len(stores) + 4}'].border = bold_border

    sheet[f'E{len(stores) + 4}'].value = df['sold_ss'].sum()
    sheet[f'E{len(stores) + 4}'].font = font
    sheet[f'E{len(stores) + 4}'].fill = PatternFill(fill_type='solid', fgColor='00C0C0C0')
    sheet[f'E{len(stores) + 4}'].alignment = Alignment(horizontal='center')
    sheet[f'E{len(stores) + 4}'].border = bold_border

    sheet[f'F{len(stores) + 4}'].value = df['stonks'].sum()
    sheet[f'F{len(stores) + 4}'].font = font
    sheet[f'F{len(stores) + 4}'].fill = PatternFill(fill_type='solid', fgColor='00C0C0C0')
    sheet[f'F{len(stores) + 4}'].alignment = Alignment(horizontal='center')
    sheet[f'F{len(stores) + 4}'].border = bold_border

    if sheet['H2'].value is None:
        plot1 = Image(plot_name_hist)
        sheet.add_image(plot1, "H2")
        plot1.height = 300
        plot1.width = 900
        sheet['H2'].value = '1'
        sheet['H2'].font = Font(color='00FFFFFF')

    if sheet[f'B{len(stores) + 5}'].value is None:
        plot2 = Image(plot_name_line)
        sheet.add_image(plot2, f'B{len(stores) + 5}')
        # plot1.height = 300
        # plot1.width = 900
        sheet[f'B{len(stores) + 5}'].value = '2'
        sheet[f'B{len(stores) + 5}'].font = Font(color='00FFFFFF')


    if sheet['H15'].value is None:
        plot3 = Image(plot_name_hist_bad_goods)
        sheet.add_image(plot3, 'H15')
        plot3.height = 300
        plot3.width = 900
        sheet['H15'].value = '3'
        sheet['H15'].font = Font(color='00FFFFFF')

    if sheet['H32'].value is None:
        plot4 = Image(plot_name_acc_hist)
        sheet.add_image(plot4, 'H32')
        plot4.height = 300
        plot4.width = 900
        sheet['H32'].value = '4'
        sheet['H32'].font = Font(color='00FFFFFF')

    if sheet['O2'].value is None:
        plot5 = Image(plot_name_pie)
        sheet.add_image(plot5, 'O2')
        # plot5.height = 300
        # plot5.width = 900
        sheet['O2'].value = '5'
        sheet['O2'].font = Font(color='00FFFFFF')

    label = Image('label.jpg')
    sheet.add_image(label, 'O34')
    label.height = 207
    label.width = 720

    wb.save('Отчет.xlsx')




def main():
    df = db_getdata("SELECT * FROM sale_day")
    df = prepare_data(df)
    plot_name_hist = group_by_art_id_top_goods(df)
    plot_name_line = group_by_date(df)
    plot_name_acc_hist = group_by_store_accum_hist(df)
    plot_name_hist_bad_goods = group_by_art_id_bad_goods(df)
    plot_name_pie = group_by_store_id_pie(df)
    load_to_excel_sheet(df, plot_name_hist, plot_name_line, plot_name_acc_hist, plot_name_hist_bad_goods, plot_name_pie)




if __name__ == '__main__':
    main()

